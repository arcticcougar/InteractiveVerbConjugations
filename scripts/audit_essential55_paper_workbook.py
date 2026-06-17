from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

import generate_essential55_paper_workbook as gen


HIGHLIGHT_RGB = "FFFF00A8"


def cell_text(cell) -> str:
    return gen.clean_text(str(cell.value or ""))


def highlighted_ranges(cell) -> list[tuple[int, int]]:
    value = cell.value
    if not isinstance(value, CellRichText):
        return []
    ranges: list[tuple[int, int]] = []
    cursor = 0
    for part in value:
        text = part.text if isinstance(part, TextBlock) else str(part)
        start = cursor
        cursor += len(text)
        font = getattr(part, "font", None)
        color = getattr(font, "color", None)
        rgb = getattr(color, "rgb", None)
        if rgb == HIGHLIGHT_RGB:
            ranges.append((start, cursor))
    return gen.merge_ranges(ranges)


def expected_value_for_offset(
    offset: int,
    verb_index: int,
    block: dict,
    verbs: list[dict],
    maps: list[dict[str, str]],
) -> str:
    if offset == 2:
        return gen.clean_text(verbs[verb_index].get("infinitive"))
    cell_key = next((key for key, row_offset in gen.FORM_ROWS.items() if row_offset == offset), "")
    if cell_key and block["filled"]:
        return maps[verb_index].get(cell_key, "")
    return ""


def expected_highlight_for_offset(
    offset: int,
    verb_index: int,
    block: dict,
    maps: list[dict[str, str]],
    regular_maps: list[dict[str, str]],
) -> list[tuple[int, int]]:
    if not block["filled"]:
        return []
    cell_key = next((key for key, row_offset in gen.FORM_ROWS.items() if row_offset == offset), "")
    if not cell_key:
        return []
    actual = maps[verb_index].get(cell_key, "")
    regular = regular_maps[verb_index].get(cell_key, "")
    return gen.irregular_ranges(actual, regular)


def audit_workbook(path: Path) -> list[str]:
    data = gen.load_course_data()
    weeks = data["weeks"]
    verbs_by_infinitive = {gen.normalize(verb["infinitive"]): verb for verb in data["verbs"]}
    errors: list[str] = []

    wb = load_workbook(path, rich_text=True, data_only=True)
    try:
        expected_sheets = [f"Week {week['week']}" for week in weeks]
        actual_sheets = wb.sheetnames
        if actual_sheets[: len(expected_sheets)] != expected_sheets:
            errors.append(f"Sheet order/name mismatch: expected first sheets {expected_sheets}, got {actual_sheets}")
        if len(actual_sheets) != len(expected_sheets):
            errors.append(f"Expected {len(expected_sheets)} sheets, found {len(actual_sheets)}")

        for week in weeks:
            sheet_name = f"Week {week['week']}"
            if sheet_name not in wb.sheetnames:
                errors.append(f"Missing sheet {sheet_name}")
                continue
            ws = wb[sheet_name]
            if ws.row_dimensions[3].height != 1.8:
                errors.append(f"{sheet_name}: row 3 height is {ws.row_dimensions[3].height}, expected 1.8")
            if ws.row_dimensions[35].height != 1.8:
                errors.append(f"{sheet_name}: row 35 height is {ws.row_dimensions[35].height}, expected 1.8")

            try:
                verbs = [verbs_by_infinitive[gen.normalize(name)] for name in week["verbs"]]
            except KeyError as exc:
                errors.append(f"{sheet_name}: missing verb data for {exc}")
                continue
            maps = [gen.canonical_cell_map(verb) for verb in verbs]
            regular_maps = [gen.regular_expected_map(verb, verbs_by_infinitive) for verb in verbs]

            for block in gen.BLOCKS:
                row0 = block["row"]
                label_col = block["label_col"]
                answer_col = block["answer_col"]
                expected_header = f"Essential 55 - Week {week['week']} {block['label']}"
                actual_header = cell_text(ws.cell(row0, label_col))
                if actual_header != expected_header:
                    errors.append(
                        f"{sheet_name} {block['label']} block: header {actual_header!r}, expected {expected_header!r}"
                    )

                for offset in range(2, 30):
                    for verb_index in range(3):
                        cell = ws.cell(row0 + offset, answer_col + verb_index)
                        expected = expected_value_for_offset(offset, verb_index, block, verbs, maps)
                        actual = cell_text(cell)
                        if actual != gen.clean_text(expected):
                            errors.append(
                                f"{sheet_name} {block['label']} {cell.coordinate}: {actual!r}, expected {expected!r}"
                            )
                        expected_ranges = expected_highlight_for_offset(
                            offset,
                            verb_index,
                            block,
                            maps,
                            regular_maps,
                        )
                        actual_ranges = highlighted_ranges(cell)
                        if actual_ranges != expected_ranges:
                            errors.append(
                                f"{sheet_name} {block['label']} {cell.coordinate}: highlight {actual_ranges}, expected {expected_ranges}"
                            )
    finally:
        wb.close()

    return errors


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit the generated Essential 55 paper workbook.")
    parser.add_argument(
        "path",
        nargs="?",
        default=str(gen.OUTPUT_PATH),
        help="Workbook path to audit",
    )
    args = parser.parse_args()
    path = Path(args.path)
    errors = audit_workbook(path)
    if errors:
        print(f"FAILED: {len(errors)} workbook audit issue(s)")
        for error in errors:
            print(error)
        raise SystemExit(1)
    print(f"PASS: audited {path.name} across 30 weeks, 4 blocks per week, 3 verbs per block.")


if __name__ == "__main__":
    main()

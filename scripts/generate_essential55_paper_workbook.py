from __future__ import annotations

import json
import re
import subprocess
import unicodedata
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT / "essential verbs paper version.xlsx"
OUTPUT_PATH = ROOT / "Essential55_Paper_Practice_Workbook.xlsx"

DEFAULT_TENSE_KEYS = ["gerund", "participle", "1", "3", "4"]
REGULAR_PATTERN_MODEL_INFINITIVES = {
    "ar": {"plain": "hablar", "reflexive": "llamarse"},
    "er": {"plain": "comer", "reflexive": "atreverse"},
    "ir": {"plain": "vivir", "reflexive": "aburrirse"},
}

BLOCKS = [
    {"row": 1, "col": 2, "label": "Reference", "filled": True},
    {"row": 1, "col": 8, "label": "Practice", "filled": False},
    {"row": 34, "col": 2, "label": "Practice", "filled": False},
    {"row": 34, "col": 8, "label": "Practice", "filled": False},
]

FORM_ROWS = {
    "h:gerund": 4,
    "h:participle": 5,
    "s:1:sg:0": 8,
    "s:1:sg:1": 9,
    "s:1:sg:2": 10,
    "s:1:pl:0": 11,
    "s:1:pl:1": 12,
    "s:1:pl:2": 13,
    "s:3:sg:0": 16,
    "s:3:sg:1": 17,
    "s:3:sg:2": 18,
    "s:3:pl:0": 19,
    "s:3:pl:1": 20,
    "s:3:pl:2": 21,
    "s:4:sg:0": 24,
    "s:4:sg:1": 25,
    "s:4:sg:2": 26,
    "s:4:pl:0": 27,
    "s:4:pl:1": 28,
    "s:4:pl:2": 29,
}


def load_course_data() -> dict:
    script = r"""
global.window = {};
require('./verbs-data.js');
const supplemental = require('./supplemental-verbs.js');
const challenges = require('./practice-challenges.js');
console.log(JSON.stringify({
  verbs: [...window.VERB_DATA, ...supplemental],
  weeks: challenges.WEEKS,
  defaultTenseKeys: challenges.DEFAULT_TENSE_KEYS
}));
"""
    result = subprocess.run(
        ["node", "-e", script],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    return json.loads(result.stdout)


def clean_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00ad", "")).strip()


def normalize(value: str | None) -> str:
    text = unicodedata.normalize("NFD", clean_text(value).lower())
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def split_infinitive(infinitive: str) -> dict:
    raw = clean_text(infinitive).lower()
    reflexive = raw.endswith("se")
    base = raw[:-2] if reflexive else raw
    ending = base[-2:] if re.search(r"(ar|er|ir)$", base) else ""
    stem = base[:-2] if ending else base
    return {"raw": raw, "reflexive": reflexive, "base": base, "ending": ending, "stem": stem}


def tense_cell_key(tense_num: int, number: str, row_index: int) -> str:
    return f"s:{tense_num}:{number}:{row_index}"


def extract_tense_number(label: str) -> int:
    match = re.match(r"\s*(\d+)", label or "")
    return int(match.group(1)) if match else 0


def ordered_tense_keys(group: dict | None) -> list[str]:
    return sorted((group or {}).keys(), key=extract_tense_number)


def canonical_cell_map(verb: dict) -> dict[str, str]:
    out = {
        "h:gerund": clean_text(verb.get("gerund")),
        "h:participle": clean_text(verb.get("past_participle")),
    }
    for key in ordered_tense_keys(verb.get("simple")):
        num = extract_tense_number(key)
        tense = verb["simple"][key] or {}
        singular = tense.get("singular") or []
        plural = tense.get("plural") or []
        for index in range(3):
            out[tense_cell_key(num, "sg", index)] = clean_text(singular[index] if index < len(singular) else "")
            out[tense_cell_key(num, "pl", index)] = clean_text(plural[index] if index < len(plural) else "")
    for key in ordered_tense_keys(verb.get("compound")):
        num = extract_tense_number(key)
        tense = verb["compound"][key] or {}
        singular = tense.get("singular") or []
        plural = tense.get("plural") or []
        for index in range(3):
            out[tense_cell_key(num, "sg", index)] = clean_text(singular[index] if index < len(singular) else "")
            out[tense_cell_key(num, "pl", index)] = clean_text(plural[index] if index < len(plural) else "")
    return out


def replace_last(text: str, search: str, replacement: str) -> str:
    index = text.rfind(search)
    if index < 0:
        return text
    return text[:index] + replacement + text[index + len(search):]


def apply_stem_mutation(stem: str, mutation_type: str) -> str:
    if mutation_type == "o_ue" and "o" in stem:
        return replace_last(stem, "o", "ue")
    if mutation_type == "e_ie" and "e" in stem:
        return replace_last(stem, "e", "ie")
    if mutation_type == "e_i" and "e" in stem:
        return replace_last(stem, "e", "i")
    if mutation_type == "u_ue" and "u" in stem:
        return replace_last(stem, "u", "ue")
    if mutation_type == "i_ie" and "i" in stem:
        return replace_last(stem, "i", "ie")
    return stem


def regular_participle(infinitive: str) -> str:
    parts = split_infinitive(infinitive)
    if parts["ending"] == "ar":
        return f"{parts['stem']}ado"
    if parts["ending"] in {"er", "ir"}:
        return f"{parts['stem']}ido"
    return parts["base"]


def transform_model_form(
    model_form: str,
    model_verb: dict,
    target_verb: dict,
    *,
    use_regular_participle: bool = False,
) -> str:
    out = clean_text(model_form)
    changed = False
    model_parts = split_infinitive(model_verb["infinitive"])
    target_parts = split_infinitive(target_verb["infinitive"])
    model_part = clean_text(model_verb.get("past_participle")) or regular_participle(model_verb["infinitive"])
    target_part = (
        regular_participle(target_verb["infinitive"])
        if use_regular_participle
        else clean_text(target_verb.get("past_participle")) or regular_participle(target_verb["infinitive"])
    )

    def replace_word(source: str, replacement: str) -> None:
        nonlocal out, changed
        if not source or source == replacement:
            return
        next_value = re.sub(rf"\b{re.escape(source)}\b", replacement, out, flags=re.IGNORECASE)
        if next_value != out:
            out = next_value
            changed = True

    replace_word(model_verb["infinitive"].lower(), target_verb["infinitive"].lower())
    replace_word(model_part.lower(), target_part.lower())

    if not changed and model_parts["stem"] and target_parts["stem"] and model_parts["stem"] != target_parts["stem"]:
        pattern = rf"(^|[^\w]){re.escape(model_parts['stem'])}(\w+)"
        next_value = re.sub(pattern, lambda m: f"{m.group(1)}{target_parts['stem']}{m.group(2)}", out, flags=re.IGNORECASE)
        if next_value != out:
            out = next_value
            changed = True

    if not changed and model_parts["stem"] and target_parts["stem"]:
        for mutation_type in ["o_ue", "e_ie", "e_i", "u_ue", "i_ie"]:
            model_stem = apply_stem_mutation(model_parts["stem"], mutation_type)
            target_stem = apply_stem_mutation(target_parts["stem"], mutation_type)
            if not model_stem or not target_stem or model_stem == model_parts["stem"]:
                continue
            pattern = rf"(^|[^\w]){re.escape(model_stem)}(\w*)"
            next_value = re.sub(pattern, lambda m: f"{m.group(1)}{target_stem}{m.group(2)}", out, flags=re.IGNORECASE)
            if next_value != out:
                out = next_value
                changed = True
    return clean_text(out)


def regular_expected_map(verb: dict, verbs_by_infinitive: dict[str, dict]) -> dict[str, str]:
    parts = split_infinitive(verb["infinitive"])
    spec = REGULAR_PATTERN_MODEL_INFINITIVES.get(parts["ending"])
    if not spec:
        return {}
    model_name = spec["reflexive"] if parts["reflexive"] else spec["plain"]
    model = verbs_by_infinitive.get(normalize(model_name)) or verbs_by_infinitive.get(normalize(spec["plain"]))
    if not model:
        return {}
    model_map = canonical_cell_map(model)
    return {
        cell_key: transform_model_form(
            model_map.get(cell_key, ""),
            model,
            verb,
            use_regular_participle=True,
        )
        for cell_key in canonical_cell_map(verb)
    }


def forms_differ(actual: str, regular: str) -> bool:
    return bool(clean_text(actual) and clean_text(regular) and normalize(actual) != normalize(regular))


def split_comparison_segments(text: str) -> list[tuple[str, int]]:
    segments = []
    cursor = 0
    for match in re.finditer(r"(\s+/\s+|;\s*|,\s*)", text or ""):
        if match.start() > cursor:
            segments.append((text[cursor:match.start()], cursor))
        cursor = match.end()
    if cursor < len(text or ""):
        segments.append((text[cursor:], cursor))
    return segments


def changed_range(actual: str, regular: str, offset: int) -> tuple[int, int] | None:
    if not forms_differ(actual, regular):
        return None
    min_len = min(len(actual), len(regular))
    start = 0
    while start < min_len and actual[start] == regular[start]:
        start += 1
    actual_end = len(actual)
    regular_end = len(regular)
    while actual_end > start and regular_end > start and actual[actual_end - 1] == regular[regular_end - 1]:
        actual_end -= 1
        regular_end -= 1
    if start >= actual_end:
        begin = max(0, min(len(actual) - 1, start - 1))
        end = min(len(actual), max(begin + 1, start + 1))
        return (offset + begin, offset + end)
    return (offset + start, offset + actual_end)


def irregular_ranges(actual: str, regular: str) -> list[tuple[int, int]]:
    actual = clean_text(actual)
    regular = clean_text(regular)
    if not forms_differ(actual, regular):
        return []
    actual_segments = split_comparison_segments(actual)
    regular_segments = split_comparison_segments(regular)
    if actual_segments and len(actual_segments) == len(regular_segments):
        ranges = []
        for (actual_text, start), (regular_text, _) in zip(actual_segments, regular_segments):
            item = changed_range(actual_text, regular_text, start)
            if item:
                ranges.append(item)
        return merge_ranges(ranges)
    item = changed_range(actual, regular, 0)
    return [item] if item else []


def merge_ranges(ranges: list[tuple[int, int]]) -> list[tuple[int, int]]:
    merged = []
    for start, end in sorted((r for r in ranges if r and r[1] > r[0])):
        if merged and start <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
        else:
            merged.append((start, end))
    return merged


def inline_font_from_cell(cell, *, color: str | None = None, bold: bool | None = None) -> InlineFont:
    font = cell.font
    return InlineFont(
        rFont=font.name or "Aptos Narrow",
        sz=font.sz or 11,
        b=font.bold if bold is None else bold,
        i=font.italic,
        color=color,
    )


def set_form_cell(cell, value: str, regular_value: str = "") -> None:
    value = clean_text(value)
    alignment = copy(cell.alignment) if cell.alignment else Alignment()
    alignment.shrinkToFit = True
    cell.alignment = alignment
    if not value:
        cell.value = None
        return
    ranges = irregular_ranges(value, regular_value)
    if not ranges:
        cell.value = value
        return

    normal_font = inline_font_from_cell(cell)
    highlight_font = inline_font_from_cell(cell, color="FFFF00A8", bold=True)
    parts = []
    cursor = 0
    for start, end in ranges:
        if start > cursor:
            parts.append(TextBlock(normal_font, value[cursor:start]))
        parts.append(TextBlock(highlight_font, value[start:end]))
        cursor = end
    if cursor < len(value):
        parts.append(TextBlock(normal_font, value[cursor:]))
    cell.value = CellRichText(*parts)


def clear_answer_cells(ws, block: dict) -> None:
    row0 = block["row"]
    col0 = block["col"]
    for offset in FORM_ROWS.values():
        for col in range(col0 + 1, col0 + 4):
            ws.cell(row0 + offset, col).value = None


def fill_block(ws, week: dict, verbs: list[dict], maps: list[dict], regular_maps: list[dict], block: dict) -> None:
    row0 = block["row"]
    col0 = block["col"]
    ws.cell(row0, col0).value = f"Essential 55 - Week {week['week']} {block['label']}"
    for index, verb in enumerate(verbs):
        set_form_cell(ws.cell(row0 + 2, col0 + 1 + index), verb["infinitive"])

    clear_answer_cells(ws, block)
    if not block["filled"]:
        return

    for index, verb_map in enumerate(maps):
        regular_map = regular_maps[index]
        for cell_key, offset in FORM_ROWS.items():
            set_form_cell(
                ws.cell(row0 + offset, col0 + 1 + index),
                verb_map.get(cell_key, ""),
                regular_map.get(cell_key, ""),
            )


def fill_week_sheet(ws, week: dict, verbs_by_infinitive: dict[str, dict]) -> None:
    verbs = [verbs_by_infinitive[normalize(name)] for name in week["verbs"]]
    maps = [canonical_cell_map(verb) for verb in verbs]
    regular_maps = [regular_expected_map(verb, verbs_by_infinitive) for verb in verbs]
    ws.title = f"Week {week['week']}"
    ws.sheet_view.showGridLines = False
    for block in BLOCKS:
        fill_block(ws, week, verbs, maps, regular_maps, block)


def main() -> None:
    data = load_course_data()
    if data.get("defaultTenseKeys") != DEFAULT_TENSE_KEYS:
        raise SystemExit(f"Unexpected default tense keys: {data.get('defaultTenseKeys')}")

    verbs_by_infinitive = {normalize(verb["infinitive"]): verb for verb in data["verbs"]}
    missing = [
        f"Week {week['week']}: {name}"
        for week in data["weeks"]
        for name in week["verbs"]
        if normalize(name) not in verbs_by_infinitive
    ]
    if missing:
        raise SystemExit("Missing weekly verbs:\n" + "\n".join(missing))

    wb = load_workbook(TEMPLATE_PATH)
    template = wb.active
    sheets = [template]
    while len(sheets) < len(data["weeks"]):
        sheets.append(wb.copy_worksheet(template))

    for ws, week in zip(sheets, data["weeks"]):
        fill_week_sheet(ws, week, verbs_by_infinitive)

    wb.active = 0
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH.relative_to(ROOT)} with {len(data['weeks'])} weekly sheets.")


if __name__ == "__main__":
    main()
